import json
import datetime
import os 

import shutil
from openpyxl import load_workbook
import psycopg2
from pymongo import MongoClient

file = "V2 - DummyData.xlsx"
role_file = "role_templates.json"
acl_file = "acl_templates.json"
dummy_image_dir = "./Images/"
base_image_dir = "../uploads/"

#post_host = ""
#post_port = 0
#post_user = ""
#post_password = ""
#post_database = ""
mongo_address = "localhost"
mongo_db = "test_lionet"
collections = ['users', 'follows', 'profiles', 'spaces', 'roles', 'space_acl', 'global_acl', 'posts']

"""
Insert users by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_users_upperRow to insert_users_lowerRow from the worksheet
with attributes between column 1 and 4 (e.g. username, email, role, hashed_password)
"""
insert_users_upperRow = 2
insert_users_lowerRow = 4
insert_users_columnLeft = 1
insert_users_columnRight = 2

"""
Insert follows by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_follows_upperRow to insert_follows_lowerRow from the worksheet
with attributes between column 1 and 2 (user, follows)
"""
insert_follows_upperRow = 2
insert_follows_lowerRow = 4
insert_follows_columnLeft = 1
insert_follows_columnRight = 2

"""
Insert profiles by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_profiles_upperRow to insert_profiles_lowerRow from the worksheet
with attributes between column 1 and 11 (e.g. user, bio, gender, ...)
"""
insert_profiles_upperRow = 2
insert_profiles_lowerRow = 4
insert_profiles_columnLeft = 1
insert_profiles_columnRight = 12

"""
Insert spaces by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_spaces_upperRow to insert_spaces_lowerRow from the worksheet
with attributes between column 1 and 2 (name, members)
"""
insert_spaces_upperRow = 2
insert_spaces_lowerRow = 3
insert_spaces_columnLeft = 1
insert_spaces_columnRight = 7

"""
Insert posts by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_posts_upperRow to insert_posts_lowerRow from the worksheet
with attributes between column 1 and 8 (e.g. author, creation_date, text, ...)
"""
insert_posts_upperRow = 2
insert_posts_lowerRow = 13
insert_posts_columnLeft = 1
insert_posts_columnRight = 13

"""
Insert roles by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_posts_upperRow to insert_posts_lowerRow from the worksheet
with attributes between column 1 and 2
"""
insert_roles_upperRow = 2
insert_roles_lowerRow = 4
insert_roles_columnLeft = 1
insert_roles_columnRight = 2

"""
Insert global acl by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_posts_upperRow to insert_posts_lowerRow from the worksheet
with attributes between column 1 and 2
"""
insert_globalacl_upperRow = 2
insert_globalacl_lowerRow = 4
insert_globalacl_columnLeft = 1
insert_globalacl_columnRight = 2

"""
Insert space acl by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_posts_upperRow to insert_posts_lowerRow from the worksheet
with attributes between column 1 and 10
"""
insert_spaceacl_upperRow = 2
insert_spaceacl_lowerRow = 6
insert_spaceacl_columnLeft = 1
insert_spaceacl_columnRight = 10


"""
Insert users to the mongo database from the dummy_data worksheet by using config values from above
Function only for clarification purposes and a more complete dummy data worksheet.
Base users are send from keycloak and not from mogno db
"""
def insert_users():
    workbook = load_workbook(filename=file)
    sheet = workbook['users']

    client = MongoClient(mongo_address)
    db = client[mongo_db]


    for value in sheet.iter_rows(min_row=insert_users_upperRow, max_row=insert_users_lowerRow, min_col=insert_users_columnLeft, max_col=insert_users_columnRight, values_only=True):
        entry = {'user': value[0]}

        db.users.insert_one(entry)
        print(f"Inserted user {value[0]} to user collection")


"""
Insert roles to the Mongo database from the dummy_data worksheet by using config values from above
"""
def insert_roles():
    workbook = load_workbook(filename=file)
    sheet = workbook['roles']

    client = MongoClient(mongo_address)
    db = client[mongo_db]


    for value in sheet.iter_rows(min_row=insert_roles_upperRow, max_row=insert_roles_lowerRow, min_col=insert_roles_columnLeft, max_col=insert_roles_columnRight, values_only=True):
        entry = {'username': value[0], 'role': value[1]}

        db.roles.insert_one(entry)
        print(f"Inserted user {value[0]} to roles collection")    

"""
Insert spaces to the Mongo database from the dummy_data worksheet by using config values from above
"""
def insert_spaces():
    workbook = load_workbook(filename=file)
    sheet = workbook['spaces']

    client = MongoClient(mongo_address)
    db = client[mongo_db]

    for value in sheet.iter_rows(min_row=insert_spaces_upperRow, max_row=insert_spaces_lowerRow, min_col=insert_spaces_columnLeft, max_col=insert_spaces_columnRight, values_only=True):
        entry = {
            'name': value[0],
            'members': [member for member in value[1].split(', ')] if value[1] != None else [],
            'admins': [admin for admin in value[2].split(', ')] if value[2] != None else [],
            'invites': [invite for invite in value[3].split(', ')] if value[3] != None else [],
            'requests': [request for request in value[4].split(', ')] if value[4] != None else [],
            "space_pic": value[5],
            'space_description': value[6]
        }

        db.spaces.insert_one(entry)
        print(f"Inserted space {value[0]} to spaces collection")    


"""
Insert follows to the Mongo database from the dummy_data worksheet by using config values from above
"""
def insert_follows():
    workbook = load_workbook(filename=file)
    sheet = workbook['follows']

    client = MongoClient(mongo_address)
    db = client[mongo_db]

    for value in sheet.iter_rows(min_row=insert_follows_upperRow, max_row=insert_follows_lowerRow, min_col=insert_follows_columnLeft, max_col=insert_follows_columnRight, values_only=True):
        entry = {'user': value[0], 'follows': []}

        for follows_user in value[1].split(', '):
            entry['follows'].append(follows_user)
        
        db.follows.insert_one(entry)
        print(f"Inserted user {value[0]} to follows collection")

"""
Insert profiles to the Mongo database from the dummy_data worksheet by using config values from above
"""
def insert_profiles():
    workbook = load_workbook(filename=file)
    sheet = workbook['profiles']

    client = MongoClient(mongo_address)
    db = client[mongo_db]

    for value in sheet.iter_rows(min_row=insert_profiles_upperRow, max_row=insert_profiles_lowerRow, min_col=insert_profiles_columnLeft, max_col=insert_profiles_columnRight, values_only=True):
        entry = {
            'user': value[0],
            'address': value[1],
            'bio': value[2],
            'birthday': str(value[3]).split(' ')[0],
            'education': [edu for edu in value[4].split(', ')] if value[4] != None else [],
            'experience': [exp for exp in value[5].split(', ')] if value[5] != None else [],
            'first_name': value[6],
            'last_name': value[7],
            'gender': value[8],
            'institution': value[9],
            'projects': [project for project in value[10].split(', ')] if value[10] != None else [],
            'profile_pic': value[11],
        }

        db.profiles.insert_one(entry)
        print(f"Inserted profile of user {value[0]} to profiles collection")

"""
Insert global acl permissions to the Mongo database from the dummy_data worksheet by using config values from above
"""
def insert_global_acl():
    workbook = load_workbook(filename=file)
    sheet = workbook['global_acl']

    client = MongoClient(mongo_address)
    db = client[mongo_db]

    for value in sheet.iter_rows(min_row=insert_globalacl_upperRow, max_row=insert_globalacl_lowerRow, min_col=insert_globalacl_columnLeft, max_col=insert_globalacl_columnRight, values_only=True):
        entry = {
            'role': value[0], 
            'create_space': eval(value[1])
        }

        db.global_acl.insert_one(entry)
        print(f"Inserted {value[0]} to global acl collection")

"""
Insert space acl permissions to the Mongo database from the dummy_data worksheet by using config values from above
"""
def insert_space_acl():
    workbook = load_workbook(filename=file)
    sheet = workbook['space_acl']

    client = MongoClient(mongo_address)
    db = client[mongo_db]

    for value in sheet.iter_rows(min_row=insert_spaceacl_upperRow, max_row=insert_spaceacl_lowerRow, min_col=insert_spaceacl_columnLeft, max_col=insert_spaceacl_columnRight, values_only=True):
        entry = {
            'role': value[0],
            'space': value[1],
            'comment': eval(value[2]),
            'join_space': eval(value[3]),
            'post': eval(value[4]),
            'read_files': eval(value[5]),
            'read_timeline': eval(value[6]),
            'read_wiki': eval(value[7]),
            'write_files': eval(value[8]),
            'write_wiki': eval(value[9]),
        }

        db.space_acl.insert_one(entry)
        print(f"Inserted {value[0]} to space acl collection")

"""
Insert posts to the Mongo database from the dummy_data worksheet by using config values from above
"""
def insert_posts():
    workbook = load_workbook(filename=file)
    sheet = workbook['posts']

    client = MongoClient(mongo_address)
    db = client[mongo_db]

    for value in sheet.iter_rows(min_row=insert_posts_upperRow, max_row=insert_posts_lowerRow, min_col=insert_posts_columnLeft, max_col=insert_posts_columnRight, values_only=True):
        entry = {
            'author': value[0],
            'creation_date': value[1],
            'text': value[2],
            'space': value[3],
            'pinned': eval(value[4]),
            'tags': value[5] if value[5] != None else "",
            'files': [file for file in value[6].split(', ')] if value[6] != None else [],
            'comments': [],
            'likers': [liker for liker in value[8].split(', ')] if value[8] != None else [],           
        }
        
        if value[7] != None:
            entry['comments'].append({
                '_id': '5fec2c0b348df9f22156cc07',
                'author': "platform_admin", 
                'creation_date': datetime.datetime.now(), 
                'text': 'Ja können wir!', 
                'pinned': True
            })           
        

        if value[8] == True:
            entry['isRepost']: eval(value[9])
            entry['repostAuthor']: value[10]
            entry['originalCreator']: value[11]
            entry['repostText']: value[12]

        db.posts.insert_one(entry)

        print(f"Inserted post from {value[0]} to post collection")

"""
Insert dummy images to the uploads folder.
Only works with write permissions for the uploads folder 
"""
def place_dummy_images():
    for f in os.listdir(os.fsencode(dummy_image_dir)):
        shutil.copyfile(dummy_image_dir + os.fsdecode(f) , base_image_dir)
    
"""
Clears the test instance of the database
"""
def clear_database(): 
    client = MongoClient(mongo_address)
    db = client[mongo_db]

    workbook = load_workbook(filename=file)

    for collection in collections:
        x = db[collection].delete_many({})
        print(f"{x.deleted_count} deleted documents in {collection}")


def clear_dummy_images():
    pass

"""
Main routine
"""
def main():
    clear_database()
    insert_users()
    insert_roles()
    insert_spaces()
    insert_follows()
    insert_profiles()
    insert_global_acl()
    insert_space_acl()
    insert_posts()    
    #place_dummy_images()

if __name__ == "__main__":
    main()
