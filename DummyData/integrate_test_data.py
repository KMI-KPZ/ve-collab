from openpyxl import load_workbook
import psycopg2
from pymongo import MongoClient

file = "DummyData.xlsx"

post_host = "localhost"
post_port = 5432
post_user = "postgres"
post_password = "admin"
post_database = "sse"
mongo_address = "localhost"

"""
Insert users by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_users_upperRow to insert_users_lowerRow from the worksheet
with attributes between column 1 and 4 (e.g. username, email, role, hashed_password)
"""
insert_users_upperRow = 3
insert_users_lowerRow = 12
insert_users_columnLeft = 1
insert_users_columnRight = 4

"""
Insert follows by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_follows_upperRow to insert_follows_lowerRow from the worksheet
with attributes between column 1 and 2 (user, follows)
"""
insert_follows_upperRow = 2
insert_follows_lowerRow = 12
insert_follows_columnLeft = 1
insert_follows_columnRight = 2

"""
Insert users by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_profiles_upperRow to insert_profiles_lowerRow from the worksheet
with attributes between column 1 and 11 (e.g. user, bio, gender, ...)
"""
insert_profiles_upperRow = 2
insert_profiles_lowerRow = 12
insert_profiles_columnLeft = 1
insert_profiles_columnRight = 11

"""
Insert users by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_spaces_upperRow to insert_spaces_lowerRow from the worksheet
with attributes between column 1 and 2 (name, members)
"""
insert_spaces_upperRow = 2
insert_spaces_lowerRow = 5
insert_spaces_columnLeft = 1
insert_spaces_columnRight = 2

"""
Insert users by selecting number of rows and corresponding columns for each attribute of the type
Insert users from insert_posts_upperRow to insert_posts_lowerRow from the worksheet
with attributes between column 1 and 8 (e.g. author, creation_date, text, ...)
"""
insert_posts_upperRow = 2
insert_posts_lowerRow = 27
insert_posts_columnLeft = 1
insert_posts_columnRight = 8


"""
Insert users to the postgres database from the dummy_data worksheet by using config values from above
"""
def insert_users():
    workbook = load_workbook(filename=file)
    workbook.sheetnames

    sheet = workbook['users']

    conn = psycopg2.connect(host=post_host, port=post_port, user=post_user, password=post_password, database=post_database)

    for value in sheet.iter_rows(min_row=insert_users_upperRow, max_row=insert_users_lowerRow, min_col=insert_users_columnLeft, max_col=insert_users_columnRight, values_only=True):
        print(value)

        cursor = conn.cursor()
        cursor.execute("INSERT INTO users (name, email, role, hashed_password) VALUES(%s, %s, %s, %s)", (value[0], value[1], value[2], value[3]))
        conn.commit()
        cursor.close()

    conn.close()


"""
Insert follows to the Mongo database from the dummy_data worksheet by using config values from above
"""
def insert_follows():
    workbook = load_workbook(filename=file)

    sheet = workbook['follows']

    client = MongoClient(mongo_address)
    db = client['lionet']


    for value in sheet.iter_rows(min_row=insert_follows_upperRow, max_row=insert_follows_lowerRow, min_col=insert_follows_columnLeft, max_col=insert_follows_columnRight, values_only=True):
        entry = {'user': value[0], 'follows': []}

        for follows_user in value[1].split(', '):
            entry['follows'].append(follows_user)

        db.follows.insert_one(entry)
        print(entry)

"""
Insert profiles to the Mongo database from the dummy_data worksheet by using config values from above
"""
def insert_profiles():
    workbook = load_workbook(filename=file)

    sheet = workbook['profiles']

    client = MongoClient(mongo_address)
    db = client['lionet']


    for value in sheet.iter_rows(min_row=insert_profiles_upperRow, max_row=insert_profiles_lowerRow, min_col=insert_profiles_columnLeft, max_col=insert_profiles_columnRight, values_only=True):
        print(value)
        entry = {
                'user': value[0],
                'bio': value[1],
                'institution': value[2],
                'projects': value[3],
                'address': value[4],
                'birthday': str(value[5]),
                'education': value[6],
                'experience': value[7],
                'first_name': value[8],
                'last_name': value[9],
                'gender': value[10]
                }

        db.profiles.insert_one(entry)
        print(entry)

"""
Insert spaces to the Mongo database from the dummy_data worksheet by using config values from above
"""
def insert_spaces():
    workbook = load_workbook(filename=file)

    sheet = workbook['spaces']

    client = MongoClient(mongo_address)
    db = client['lionet']

    for value in sheet.iter_rows(min_row=insert_spaces_upperRow, max_row=insert_spaces_lowerRow, min_col=insert_spaces_columnLeft, max_col=insert_spaces_columnRight, values_only=True):
        print(value)
        entry = {'name':value[0], 'members':[]}

        for member in value[1].split(', '):
            entry['members'].append(member)

        db.spaces.insert_one(entry)
        print(entry)

"""
Insert posts to the Mongo database from the dummy_data worksheet by using config values from above
"""
def insert_posts():
    workbook = load_workbook(filename=file)

    sheet = workbook['posts']

    client = MongoClient(mongo_address)
    db = client['lionet']


    for value in sheet.iter_rows(min_row=insert_posts_upperRow, max_row=insert_posts_lowerRow, min_col=insert_posts_columnLeft, max_col=insert_posts_columnRight, values_only=True):
        #print(value)
        entry = {
                'author': value[0],
                'creation_date': value[1],
                'text': value[2],
                'space': value[3],
                'tags': "",
                'files': [],
                'comments': [],
                'likes': []
                }

        if value[7] != None:
            for liker in value[7].split(', '):
                entry['likes'].append(liker)

        try:
            db.posts.insert_one(entry)
            print("Done")
        except PyMongoError:
            print(PyMongoError)
        print(entry)


"""
Deletes all posts by the corresponding author written by users from the worksheet
"""
def delete_all_corresponding_posts():
    client = MongoClient(mongo_address)
    db = client['lionet']

    workbook = load_workbook(filename=file)
    sheet = workbook['users']

    for value in sheet.iter_rows(min_row=insert_users_upperRow, max_row=insert_users_lowerRow, min_col=1, max_col=1, values_only=True):
        print(value[0])
        myquery = {'author' : value[0]}
        x = db.posts.delete_many(myquery)
        print(x.deleted_count, " documents deleted.")

"""
Main routine
"""
def main():
    """
    print("Main Excecuted")
    print("PostgreSQL:")
    post_host = input("Enter host address:")
    post_port = input("Enter port:")
    post_user = input("Enter username:")
    post_password = input("Enter password:")
    post_database = "sse"
    print("MongoDB:")
    mongo_address = input("Enter host address:")
    print("-----------------------------------------------------")
    todo = input("What do you want to do?")
    """
    #insert_users()
    insert_follows()
    insert_profiles()
    insert_spaces()
    #delete_all_corresponding_posts()
    insert_posts()
    #insert_posts()
main()
